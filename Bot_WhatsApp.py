import openpyxl
import requests
import json
import time

run = True
groups = {"Uber" : '120363028553822281@g.us', "Bot_Teste" : '120363145249360901@g.us'}

def cadastrar(numero, faltas, dias):
    dados = numero, faltas, dias
    for row in planilha.iter_rows(min_row=2, values_only=True):
        if numero == row[3]:
            break
    planilha.append(dados)

def send_message(contato, message):
    url = "https://api.ultramsg.com/instance56301/messages/chat"
    payload = f"token=779mjgh81bovkixh&to={contato}&body={message}"
    payload = payload.encode('utf8').decode('iso-8859-1')
    headers = {'content-type': 'application/x-www-form-urlencoded'}

    response = requests.request("POST", url, data=payload, headers=headers)

def read_message(contato):
    url = "https://api.ultramsg.com/instance56301/chats/messages"
    
    querystring = {
        "token": "779mjgh81bovkixh",
        "chatId": contato,
        "limit": 1
    }

    headers = {'content-type': 'application/x-www-form-urlencoded'}
    response = requests.request("GET", url, headers=headers, params=querystring)
    messages_read = response.text
    dados = json.loads(messages_read)

    remetente = [dados[0]["from"], dados[0]["body"]]
    
    return remetente

try:
    planilha_wb = openpyxl.load_workbook("Loja.xlsx")
except FileNotFoundError:
    planilha_wb = openpyxl.Workbook()
    planilha = planilha_wb.active
    planilha.title = "Clientes"
    planilha["A1"] = "Julho"
    planilha["A2"] = "Número"
    planilha["B2"] = "Faltas"
    planilha["C2"] = "Dias úteis"
planilha = planilha_wb.active

while run:
    tempo_inicial = time.time()
    contato = groups["Uber"]
    remetente = read_message(contato)

    if remetente[1].upper() == "UBER":
        msg1 = "Escolha uma das opções abaixo \n1- PAGAR"
        send_message(contato, msg1)
        
        opcoes = ["1"]
        passo1 = read_message(contato)

        while passo1[1] not in opcoes:
            passo1 = read_message(contato)
            if time.time() - tempo_inicial >= 30:
                send_message(contato, "TEMPO DE RESPOSTA ACABOU, FIM DA SESSÃO")
                break
            time.sleep(1)

        if remetente[0] == passo1[0]: 
            if passo1[1] == "1":
                send_message(contato, "PIX: marcoscorreademelo@gmail.com")
    
    time.sleep(3)